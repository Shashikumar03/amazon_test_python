# import openpyxl
# from openpyxl.reader.excel import load_workbook
#
#
from openpyxl.reader.excel import load_workbook


from openpyxl import load_workbook

def get_data():
    workbook = load_workbook(filename="/home/shashi/PycharmProjects/pythonSelenium/resource/login_data.xlsx")
    sheet = workbook["Sheet1"]

    data = []
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    for r in range(2, total_rows + 1):
        row_list = []
        for c in range(1, total_cols + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        data.append(row_list)

    return data

